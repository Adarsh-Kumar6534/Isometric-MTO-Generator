from fastapi import FastAPI, UploadFile, File, HTTPException, BackgroundTasks
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import Response
from typing import Dict, Any
import uuid
import time
import io
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill

from pipeline import run_pipeline

app = FastAPI(title="Isometric MTO API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# In-memory store for jobs
jobs: Dict[str, Dict[str, Any]] = {}

@app.get("/api/health")
def health_check():
    return {"status": "ok"}

async def process_job(job_id: str, file_bytes: bytes, file_ext: str):
    async def update_log(tag: str, msg: str):
        jobs[job_id]["logs"].append({"tag": tag, "msg": msg})
        
    try:
        mto = await run_pipeline(job_id, file_bytes, file_ext, update_log)
        jobs[job_id]["status"] = "completed"
        jobs[job_id]["result"] = mto.dict()
    except Exception as e:
        jobs[job_id]["status"] = "failed"
        await update_log("warn", f"[error] job failed completely: {str(e)}")

@app.post("/api/upload")
async def upload_file(background_tasks: BackgroundTasks, file: UploadFile = File(...)):
    if file.content_type not in ["image/png", "image/jpeg", "application/pdf"]:
        raise HTTPException(status_code=400, detail="Unsupported file type")
    
    file_bytes = await file.read()
    if len(file_bytes) > 20 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File too large")
        
    job_id = str(uuid.uuid4())
    ext = "pdf" if file.content_type == "application/pdf" else "image"
    
    jobs[job_id] = {
        "status": "processing",
        "logs": [],
        "result": None
    }
    
    background_tasks.add_task(process_job, job_id, file_bytes, ext)
    return {"job_id": job_id}

@app.get("/api/status/{job_id}")
def get_status(job_id: str):
    if job_id not in jobs:
        raise HTTPException(status_code=404, detail="Job not found")
    
    job = jobs[job_id]
    return {
        "status": job["status"],
        "logs": job["logs"]
    }

@app.get("/api/mto/{job_id}")
def get_mto(job_id: str):
    if job_id not in jobs:
        raise HTTPException(status_code=404, detail="Job not found")
        
    job = jobs[job_id]
    if job["status"] == "processing":
        return {"status": "processing"}
    if job["status"] == "failed":
        raise HTTPException(status_code=500, detail="Job failed")
        
    return {"status": "completed", "data": job["result"]}

@app.get("/api/mto/{job_id}/csv")
def get_mto_csv(job_id: str):
    if job_id not in jobs:
        raise HTTPException(status_code=404, detail="Job not found")
        
    job = jobs[job_id]
    if job["status"] != "completed" or not job["result"]:
        raise HTTPException(status_code=400, detail="MTO not ready")
        
    mto_data = job["result"]
    items = mto_data.get("items", [])
    
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Header
    writer.writerow(["#", "Category", "Description", "Size", "Sched/Class", "Material", "End", "Qty", "Unit", "Length (m)", "Confidence", "Remarks"])
    
    for item in items:
        writer.writerow([
            item.get("item_no"),
            item.get("category"),
            item.get("description"),
            item.get("size_nps"),
            item.get("schedule_rating"),
            item.get("material_spec"),
            item.get("end_type"),
            item.get("quantity"),
            item.get("unit"),
            item.get("length_m", ""),
            item.get("confidence"),
            item.get("remarks", "")
        ])
        
    return Response(content=output.getvalue(), media_type="text/csv", headers={"Content-Disposition": f"attachment; filename=mto_{job_id}.csv"})

@app.get("/api/mto/{job_id}/excel")
def get_mto_excel(job_id: str):
    if job_id not in jobs:
        raise HTTPException(status_code=404, detail="Job not found")
        
    job = jobs[job_id]
    if job["status"] != "completed" or not job["result"]:
        raise HTTPException(status_code=400, detail="MTO not ready")
        
    mto_data = job["result"]
    items = mto_data.get("items", [])
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Material Take-Off"
    
    headers = ["#", "Category", "Description", "Size", "Sched/Class", "Material", "End", "Qty", "Unit", "Length (m)", "Confidence", "Remarks"]
    ws.append(headers)
    
    # Style headers
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        
    # Add data
    for item in items:
        ws.append([
            item.get("item_no"),
            item.get("category"),
            item.get("description"),
            item.get("size_nps"),
            item.get("schedule_rating"),
            item.get("material_spec"),
            item.get("end_type"),
            item.get("quantity"),
            item.get("unit"),
            item.get("length_m", ""),
            item.get("confidence"),
            item.get("remarks", "")
        ])
        
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    output = io.BytesIO()
    wb.save(output)
    
    return Response(
        content=output.getvalue(), 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
        headers={"Content-Disposition": f"attachment; filename=mto_{job_id}.xlsx"}
    )
